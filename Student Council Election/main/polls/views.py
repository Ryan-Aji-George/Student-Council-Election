from django.template import loader
from django.http import HttpResponse, HttpResponseRedirect
from .models import Post, Candidate
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout

from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from .models import Post, Candidate, Vote
from users.models import Profile  # Import Profile model


@login_required
def vote_all(request):
    # Get all non-house captain posts
    all_posts = Post.objects.filter(is_house_captain=False).order_by('order')

    # Get posts the user has already voted for
    voted_post_ids = Vote.objects.filter(user=request.user).values_list('post_id', flat=True)

    # Filter out already voted posts
    unvoted_posts = all_posts.exclude(id__in=voted_post_ids)

    # If all posts are voted, redirect to home
    if not unvoted_posts.exists():
        logout(request)
        return redirect('polls:al_success')

    if request.method == "POST":
        error_message = None
        selected_candidates = {}

        # Ensure a candidate is selected for each remaining post
        for post in unvoted_posts:
            candidate_id = request.POST.get(f'post_{post.id}')
            if not candidate_id:
                error_message = "You must select a candidate for every post."
                break
            selected_candidates[post.id] = candidate_id

        if error_message:
            return render(request, 'polls/VoteAll.html', {'latest_post_list': unvoted_posts, 'error_message': error_message})

        # Process votes
        for post_id, candidate_id in selected_candidates.items():
            selected_candidate = Candidate.objects.get(pk=candidate_id)
            selected_candidate.votes += 1
            selected_candidate.save()

            # Record the vote
            Vote.objects.create(user=request.user, post_id=post_id)

        logout(request)  # Log out the user after voting
        return HttpResponseRedirect(reverse('polls:success_vote_all'))

    return render(request, 'polls/VoteAll.html', {'latest_post_list': unvoted_posts})

@login_required
def general_vote(request):
    user_profile = Profile.objects.get(user=request.user)  # Get user's house

    # Get all posts the user hasn't voted for
    unvoted_posts = Post.objects.filter(is_house_captain=False).exclude(
        id__in=Vote.objects.filter(user=request.user).values_list('post_id', flat=True)
    ).order_by('order')

    # Include house-specific posts the user hasn't voted for
    house_posts = Post.objects.filter(house=user_profile.house).exclude(
        id__in=Vote.objects.filter(user=request.user).values_list('post_id', flat=True)
    ).order_by('order')

    # Combine general and house-specific posts
    all_posts = unvoted_posts | house_posts

    # If all posts have been voted on, redirect to home
    if not all_posts.exists():
        logout(request)
        return redirect('polls:al_success')

    if request.method == "POST":
        error_message = None
        selected_candidates = {}

        for post in all_posts:
            candidate_id = request.POST.get(f'post_{post.id}')
            if not candidate_id:
                error_message = "You must select a candidate for every post."
                break
            selected_candidates[post.id] = candidate_id

        if error_message:
            return render(request, 'polls/GeneralVote.html', {'latest_post_list': all_posts, 'error_message': error_message})

        for post_id, candidate_id in selected_candidates.items():
            selected_candidate = Candidate.objects.get(pk=candidate_id)
            selected_candidate.votes += 1
            selected_candidate.save()

            # Record the vote
            Vote.objects.create(user=request.user, post_id=post_id)

        logout(request)
        return HttpResponseRedirect(reverse('polls:success_vote_all'))

    return render(request, 'polls/GeneralVote.html', {'latest_post_list': all_posts})


def get_available_houses():
    """Returns a list of houses that have at least one post, excluding None and 'Not a House Captain'."""
    houses = Post.objects.exclude(house__isnull=True).exclude(house="None").values_list('house', flat=True).distinct()

    def get_none_posts():
        """Returns posts that do not belong to a house."""
        return Post.objects.filter(models.Q(house__isnull=True) | models.Q(house="Not a House Captain")).order_by(
            'order')
    return [house for house in houses if Post.objects.filter(house=house).exists()]

def index(request):
    latest_post_list = Post.objects.order_by('order')
    context = {
        'latest_post_list': latest_post_list,

    }
    return render(request, 'polls/index.html', context)


def select_all(request):
    house_captain_posts1 = Post.objects.filter(is_house_captain=False).order_by('order')  # Only house captain posts
    house_captain_posts2 = Post.objects.filter(is_house_captain=True).order_by('order')
    available_houses = get_available_houses()
    return render(request, 'polls/select.html', {'latest_post_list1': house_captain_posts1, 'latest_post_list2': house_captain_posts2, 'available_houses': available_houses,})

@login_required
def select(request):
    user_profile = Profile.objects.get(user=request.user)  # Get user's house

    house_captain_posts1 = Post.objects.filter(is_house_captain=False).order_by('order')  # General posts
    house_captain_posts2 = Post.objects.filter(is_house_captain=True, house=user_profile.house).order_by('order')  # House captain posts of user's house

    available_houses = [user_profile.house] if Post.objects.filter(house=user_profile.house).exists() else []

    return render(request, 'polls/select.html', {
        'latest_post_list1': house_captain_posts1,
        'latest_post_list2': house_captain_posts2,
        'available_houses': available_houses,
    })
def success(request):
    return render(request, 'polls/success.html')

def al_success(request):
    return render(request, 'polls/al_success.html')

def al_success2(request):
    return render(request, 'polls/al_success-2.html')

def success_vote_all(request):
    available_houses = get_available_houses()
    return render(request, 'polls/success_vote_all.html', {'available_houses': available_houses,})

from .models import Vote  # Import Vote model

@login_required
def detail(request, post_id):
    post = get_object_or_404(Post, pk=post_id)

    # Check if the user has already voted for this post
    if Vote.objects.filter(user=request.user, post=post).exists():
        logout(request)
        return redirect('polls:al_success-2')  # Redirect to home page

    return render(request, 'polls/detail.html', {'post': post})


@login_required
def vote(request, post_id):
    post = get_object_or_404(Post, pk=post_id)

    # Prevent multiple votes
    if Vote.objects.filter(user=request.user, post=post).exists():
        logout(request)
        return redirect('polls:al_success-2')  # Redirect if user already voted

    try:
        selected_candidate = post.candidate_set.get(pk=request.POST['candidate'])
    except (KeyError, Candidate.DoesNotExist):
        return render(request, 'polls/detail.html', {
            'post': post,
            'error_message': "You didn't select a candidate.",
        })
    else:
        selected_candidate.votes += 1
        selected_candidate.save()

        # Record the vote
        Vote.objects.create(user=request.user, post=post)

        logout(request)  # Log out user after voting
        return HttpResponseRedirect(reverse('polls:success'))



from django.http import HttpResponse, HttpResponseForbidden
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from .models import Post, Candidate
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import redirect
from django.contrib.admin.views.decorators import staff_member_required

def superuser_required(view_func):
    """Custom decorator to restrict access to superusers and redirect unauthorized users."""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_superuser:
            return redirect('/admin/login/?next=' + request.path)  # Redirect to admin login
        return view_func(request, *args, **kwargs)
    return wrapper

@superuser_required
def download_results(request):
    # Create the HTTP response with PDF content type
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="poll_results.pdf"'

    # Create a PDF document
    pdf = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    margin = 50
    line_height = 20
    y_position = height - margin  # Start position for writing

    def check_page_space():
        """Check if there's enough space for more content, else create a new page."""
        nonlocal y_position
        if y_position < margin:
            pdf.showPage()  # Create a new page
            pdf.setFont("Helvetica", 12)  # Reset font
            y_position = height - margin  # Reset y_position

    pdf.setTitle("Poll Results")
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(250, y_position, "AC School")
    y_position -= 15
    pdf.drawString(167, y_position, "Student Council Election - Results")
    y_position -= 30

    posts = Post.objects.all().order_by('order')

    for post in posts:
        check_page_space()  # Ensure there's room for post title

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(50, y_position, f"{post.post_title}")
        y_position -= line_height

        candidates = post.candidate_set.order_by('-votes')
        winner = candidates.first() if candidates.exists() else None

        for candidate in candidates:
            check_page_space()  # Ensure there's room for candidate details

            if candidate == winner:
                pdf.setFont("Helvetica-Bold", 12)  # Make winner bold
            else:
                pdf.setFont("Helvetica", 12)

            pdf.drawString(70, y_position, f"{candidate.candidate_name} - {candidate.votes} votes")
            y_position -= line_height

        y_position -= 10  # Extra space between posts

    pdf.save()
    return response

@superuser_required
def results(request):
    posts = Post.objects.all().order_by('order')
    post_results = []

    for post in posts:
        candidates = post.candidate_set.order_by('-votes')  # Order by votes descending
        winner = candidates.first() if candidates else None  # The candidate with the most votes
        post_results.append({'post': post, 'candidates': candidates, 'winner': winner})

    return render(request, 'polls/results.html', {'post_results': post_results})


from .models import Vote  # Import the Vote model

@login_required
def house_vote(request, house_name):
    # Get posts related to the house
    house_posts = Post.objects.filter(house=house_name).order_by('order')

    # Filter out posts the user has already voted for
    unvoted_posts = house_posts.exclude(id__in=Vote.objects.filter(user=request.user).values_list('post', flat=True))

    # If the user has already voted for all house posts, redirect them to home
    if not unvoted_posts.exists():
        logout(request)
        return redirect('polls:al_success')

    if request.method == "POST":
        error_message = None
        selected_candidates = {}

        for post in unvoted_posts:  # Iterate only through unvoted posts
            candidate_id = request.POST.get(f'post_{post.id}')
            if not candidate_id:
                error_message = "You must select a candidate for every post."
                break
            selected_candidates[post.id] = candidate_id

        if error_message:
            return render(request, 'polls/house_vote.html', {
                'house_posts': unvoted_posts,
                'house_name': house_name,
                'error_message': error_message
            })

        # Process the votes
        for post_id, candidate_id in selected_candidates.items():
            selected_candidate = Candidate.objects.get(pk=candidate_id)
            selected_candidate.votes += 1
            selected_candidate.save()

            # Record the vote in the Vote model
            Vote.objects.create(user=request.user, post_id=post_id)

        logout(request)  # Log the user out after voting
        return HttpResponseRedirect(reverse('polls:success'))

    return render(request, 'polls/house_vote.html', {'house_posts': unvoted_posts, 'house_name': house_name})


    return render(request, 'polls/house_vote.html', {'house_posts': house_posts, 'house_name': house_name})





from django.http import HttpResponse
from django.contrib.auth.models import User
from users.models import Profile  # Adjust import based on your app structure
from openpyxl import Workbook
from io import BytesIO
import re

# Assuming superuser_required decorator is defined elsewhere
from polls.views import superuser_required  # Adjust import based on your structure

@superuser_required
def download_voters_page(request):
    # Check if the request is for the Excel file
    if 'download' in request.GET:
        # Get unique classes
        unique_classes = Profile.objects.values_list('voter_class', flat=True).distinct()

        # Define sorting key for classes
        def class_key(cls):
            if cls == "STAFF":
                return (100, "")
            else:
                match = re.match(r'(\d+)([A-Za-z])', cls)
                if match:
                    number = int(match.group(1))
                    letter = match.group(2).upper()
                    return (number, letter)
                else:
                    return (99, cls)

        sorted_classes = sorted(unique_classes, key=class_key)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Voters List"

        for cls in sorted_classes:
            # Write class header
            ws.append([f"{cls}"])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)

            # Write table headers
            
            #ws.append(["Voter Name", "Username", "Password", "Class"])

            # Get and sort profiles in this class
            class_profiles = Profile.objects.filter(voter_class=cls).order_by('voter_name')
            for profile in class_profiles:
                ws.append([profile.voter_name, profile.voter_class, 'Username : ' + profile.user.username, 'Password : ' + profile.user.username])

            # Add a blank row
            ws.append([])

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="voters-list.xlsx"'
        return response

    # Return the HTML page
    return render(request, 'polls/voters_list.html')




# polls/views.py

from openpyxl import Workbook
from io import BytesIO
from django.http import HttpResponse

@superuser_required
def download_results_excel(request):
    """
    Export all poll results into a single sheet:
    - Post title in column A (bold, not merged)
    - One-time header: "Candidate" and "Votes" in row 1
    - Each post's candidates listed below their post title
    - Blank row between posts
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Poll Results"

    # Write the single header row at the top
    ws.cell(row=1, column=1, value="Candidate").font = ws.cell(row=1, column=1).font.copy(bold=True)
    ws.cell(row=1, column=2, value="Votes").font = ws.cell(row=1, column=2).font.copy(bold=True)

    row = 3 # Start data after header
    for post in Post.objects.all().order_by('order'):
        # Post title (bold)
        ws.cell(row=row, column=1, value=post.post_title).font = ws.cell(row=row, column=1).font.copy(bold=True)
        row += 1

        # List each candidate sorted by ID
        for cand in post.candidate_set.order_by('id'):
            ws.cell(row=row, column=1, value=cand.candidate_name)
            ws.cell(row=row, column=2, value=cand.votes)
            row += 1

        # Blank spacer row between posts
        row += 1

    # Write to BytesIO and return response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = 'attachment; filename="poll_results.xlsx"'
    return resp


